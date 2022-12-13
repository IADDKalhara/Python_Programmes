#####   ADD 10% DISCOUNT TO SALE VALUE   #####

# Takes a spreadsheet and adds a discount to prices.
# Saves the new price in a new column.
# Create a Barchart

import openpyxl as xl
from openpyxl import BarChart, Reference


def process_workbook(filename):

    # selecting necessary file to work
    data_file = xl.load_workbook(filename)
    sheet1 = data_file["Sheet1"]


    for row in range(2, sheet1.max_row + 1):
        cell = sheet1.cell(row, 3)
        new_price = cell.value * 0.9

        # saving new value in a new cell
        new_price_cell = sheet1.cell(row, 5)
        new_price_cell.value = new_price

    # selecting data cells for making a chart
    values = Reference(sheet1, min_row = 1, max_row = 4, min_col = 4, max_col = 4)

    chart = BarChart()
    chart.add_data(values)
    sheet1.add_chart(chart, "e2")

    # saving as a new file
    data_file.save(filename)


